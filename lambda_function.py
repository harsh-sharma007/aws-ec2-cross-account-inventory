import boto3
import pandas as pd
import io
import os
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. Global Configurations & Validation
# ==========================================
AWS_REGION_SES = os.environ.get('SES_REGION', 'ap-south-1')
CROSS_ACCOUNT_ROLE_NAME = os.environ.get('ROLE_NAME', 'EC2InventoryRole')

SENDER_EMAIL = os.environ.get('SENDER_EMAIL')
raw_recipients = os.environ.get('RECIPIENT_EMAILS')
raw_targets = os.environ.get('TARGET_ACCOUNTS')

if not SENDER_EMAIL or not raw_recipients or not raw_targets:
    raise ValueError("CRITICAL CONFIGURATION ERROR: Environment variables are missing.")

RECIPIENT_EMAILS = [email.strip() for email in raw_recipients.split(',')]
TARGET_ACCOUNTS = [acc.strip() for acc in raw_targets.split(',')]

# ==========================================
# 2. AWS Authentication & Metrics Helper
# ==========================================
def assume_role(account_id):
    sts_client = boto3.client('sts', region_name='ap-south-1')
    role_arn = f"arn:aws:iam::{account_id}:role/{CROSS_ACCOUNT_ROLE_NAME}"
    try:
        response = sts_client.assume_role(RoleArn=role_arn, RoleSessionName="EC2InventorySession")
        credentials = response['Credentials']
        return boto3.Session(
            aws_access_key_id=credentials['AccessKeyId'],
            aws_secret_access_key=credentials['SecretAccessKey'],
            aws_session_token=credentials['SessionToken']
        )
    except Exception as e:
        print(f"Failed to assume role in account {account_id}. Error: {str(e)}")
        return None

def get_cpu_utilization(cw_client, instance_id):
    end_time = datetime.utcnow()
    start_time = end_time - timedelta(days=30)
    response = cw_client.get_metric_statistics(
        Namespace='AWS/EC2', MetricName='CPUUtilization',
        Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}],
        StartTime=start_time, EndTime=end_time, Period=86400, Statistics=['Average']
    )
    datapoints = response.get('Datapoints', [])
    if not datapoints: return 0.0
    return sum(dp['Average'] for dp in datapoints) / len(datapoints)

# ==========================================
# 3. Core Inventory Logic (Read-Only)
# ==========================================
def gather_inventory_for_account(session, account_id):
    ec2_main = session.client('ec2', region_name='ap-south-1')
    try:
        regions = [r['RegionName'] for r in ec2_main.describe_regions()['Regions']]
    except Exception as e:
        print(f"Cannot describe regions in {account_id}. Skipping.")
        return [], []
    
    inventory = []
    unnamed_asg_inventory = []
    
    for region in regions:
        ec2 = session.client('ec2', region_name=region)
        cw = session.client('cloudwatch', region_name=region)
        
        paginator = ec2.get_paginator('describe_instances')
        for page in paginator.paginate():
            for reservation in page['Reservations']:
                for inst in reservation['Instances']:
                    inst_id = inst['InstanceId']
                    tags = inst.get('Tags', [])
                    
                    asg_name = next((tag['Value'] for tag in tags if tag['Key'] == 'aws:autoscaling:groupName'), "None")
                    
                    current_name = next((tag['Value'] for tag in tags if tag['Key'] == 'Name'), "").strip()
                    
                    # Read-only evaluation block
                    if not current_name and asg_name != "None":
                        name = ""
                        unnamed_asg_inventory.append({
                            'InstanceID': inst_id,
                            'Name': name,
                            'PrivateIP': inst.get('PrivateIpAddress', 'N/A'),
                            'State': inst['State']['Name'],
                            'Region': region,
                            'InstanceType': inst['InstanceType'],
                            'ASG_Name': asg_name
                        })
                    else:
                        name = current_name if current_name else "N/A"
                    
                    sg_list = []
                    for sg in inst.get('SecurityGroups', []):
                        sg_name = sg.get('GroupName', 'Unknown')
                        sg_id = sg.get('GroupId', 'Unknown')
                        sg_list.append(f"{sg_name} ({sg_id})")
                    sg_string = "; ".join(sg_list) if sg_list else "None"
                    
                    ebs_mappings = inst.get('BlockDeviceMappings', [])
                    ebs_list = [f"{dev.get('DeviceName', 'Unknown')}:{dev.get('Ebs', {}).get('VolumeId', 'N/A')}" for dev in ebs_mappings]
                    ebs_string = "; ".join(ebs_list) if ebs_list else "None"
                    
                    all_tags = [f"{t['Key']}={t['Value']}" for t in tags]
                    tags_string = "; ".join(all_tags) if all_tags else "None"
                    
                    inventory.append({
                        'InstanceName': name,
                        'InstanceID': inst_id,
                        'State': inst['State']['Name'],
                        'Region': region,
                        'InstanceType': inst['InstanceType'],
                        'Platform': inst.get('PlatformDetails', 'Linux/UNIX'),
                        'PrivateIP': inst.get('PrivateIpAddress', 'N/A'),
                        'PublicIP': inst.get('PublicIpAddress', 'N/A'),
                        'SecurityGroups': sg_string,
                        'ASG_Membership': asg_name,
                        'EBS_Volumes': ebs_string,
                        'Total_EBS_Count': len(ebs_mappings),
                        'Tags': tags_string,
                        'AMI': inst['ImageId'],
                        'Avg_CPU_30Days': f"{round(get_cpu_utilization(cw, inst_id), 2)}%"
                    })
                    
    return inventory, unnamed_asg_inventory

# ==========================================
# 4. Lambda Handler & Execution Routing
# ==========================================
def lambda_handler(event, context):
    account_data_map = {}
    account_unnamed_map = {}
    
    sts_client = boto3.client('sts', region_name='ap-south-1')
    local_account_id = sts_client.get_caller_identity()['Account']
    
    for account in TARGET_ACCOUNTS:
        if account == local_account_id:
            session = boto3.Session(region_name='ap-south-1')
        else:
            session = assume_role(account)
            
        if session:
            inv_data, unnamed_data = gather_inventory_for_account(session, account)
            if inv_data: 
                account_data_map[account] = inv_data
            if unnamed_data:
                account_unnamed_map[account] = unnamed_data
                
    if not account_data_map:
        return {"statusCode": 200, "body": "No instances found."}
        
    excel_buffer = io.BytesIO()
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
                         
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        for account_id, inventory_list in account_data_map.items():
            df = pd.DataFrame(inventory_list)
            df.to_excel(writer, sheet_name=str(account_id), index=False, startrow=1)
            
            worksheet = writer.sheets[str(account_id)]
            total_columns = len(df.columns)
            
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f"AWS Account ID: {account_id}"
            title_cell.font = Font(bold=True)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
            
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=total_columns):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = center_alignment
            
            # Secondary Table Injection
            unnamed_list = account_unnamed_map.get(account_id, [])
            start_row_unnamed = -1
            
            if unnamed_list:
                start_row_unnamed = worksheet.max_row + 3
                
                sub_title_cell = worksheet.cell(row=start_row_unnamed, column=1)
                sub_title_cell.value = "Unnamed ASG Instances (Manual Action Required)"
                sub_title_cell.font = Font(bold=True)
                sub_title_cell.alignment = center_alignment
                
                # Extended the merge to cover 7 columns
                worksheet.merge_cells(start_row=start_row_unnamed, start_column=1, end_row=start_row_unnamed, end_column=7)
                
                headers = ["InstanceID", "Name", "PrivateIP", "State", "Region", "InstanceType", "ASG_Name"]
                for col_idx, header in enumerate(headers, 1):
                    header_cell = worksheet.cell(row=start_row_unnamed + 1, column=col_idx, value=header)
                    header_cell.font = Font(bold=True)
                    header_cell.border = thin_border
                    header_cell.alignment = center_alignment
                    
                current_row = start_row_unnamed + 2
                for item in unnamed_list:
                    cells = [
                        worksheet.cell(row=current_row, column=1, value=item['InstanceID']),
                        worksheet.cell(row=current_row, column=2, value=item['Name']),
                        worksheet.cell(row=current_row, column=3, value=item['PrivateIP']),
                        worksheet.cell(row=current_row, column=4, value=item['State']),
                        worksheet.cell(row=current_row, column=5, value=item['Region']),
                        worksheet.cell(row=current_row, column=6, value=item['InstanceType']),
                        worksheet.cell(row=current_row, column=7, value=item['ASG_Name'])
                    ]
                    
                    for c in cells:
                        c.border = thin_border
                        c.alignment = center_alignment
                    current_row += 1
            
            # Auto-adjust Column Widths Dynamically
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1 or cell.row == start_row_unnamed:
                        continue
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    total_instances = sum(len(v) for v in account_data_map.values())
    current_time_utc = datetime.utcnow()
    
    msg = MIMEMultipart('mixed')
    current_month = current_time_utc.strftime('%B')
    msg['Subject'] = f"AWS EC2 Inventory - {current_month}"
    
    msg['From'] = SENDER_EMAIL
    msg['To'] = ", ".join(RECIPIENT_EMAILS)
    
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #333333; line-height: 1.6;">
        <h2 style="color: #232F3E; border-bottom: 2px solid #FF9900; padding-bottom: 10px;">
            AWS EC2 Infrastructure Report
        </h2>
        <p>Hello,</p>
        <p>Please find attached your automated multi-account EC2 inventory report.</p>
        
        <table style="border-collapse: collapse; width: 100%; max-width: 600px; margin-top: 20px; border: 1px solid #dddddd;">
            <tr style="background-color: #f4f4f4;">
                <td style="padding: 12px; border: 1px solid #dddddd;"><strong>Accounts Scanned</strong></td>
                <td style="padding: 12px; border: 1px solid #dddddd;">{', '.join(TARGET_ACCOUNTS)}</td>
            </tr>
            <tr>
                <td style="padding: 12px; border: 1px solid #dddddd;"><strong>Total EC2 Instances</strong></td>
                <td style="padding: 12px; border: 1px solid #dddddd;"><strong>{total_instances}</strong></td>
            </tr>
            <tr style="background-color: #f4f4f4;">
                <td style="padding: 12px; border: 1px solid #dddddd;"><strong>Generated At (UTC)</strong></td>
                <td style="padding: 12px; border: 1px solid #dddddd;">{current_time_utc.strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
        </table>
        
        <p style="margin-top: 30px; font-size: 0.85em; color: #888888;">
            <em>This is an automated message generated by AWS Lambda. Please do not reply.</em>
        </p>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(html_body, 'html'))
    
    filename = f"EC2_Inventory_{current_time_utc.strftime('%Y%m%d')}.xlsx"
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(excel_buffer.getvalue())
    encoders.encode_base64(attachment)
    attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
    msg.attach(attachment)
    
    ses = boto3.client('ses', region_name=AWS_REGION_SES)
    ses.send_raw_email(Source=msg['From'], Destinations=RECIPIENT_EMAILS, RawMessage={'Data': msg.as_string()})
    
    return {"statusCode": 200, "body": "Inventory dispatched."}
